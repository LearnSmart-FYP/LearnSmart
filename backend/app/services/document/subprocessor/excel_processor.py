import base64
import shutil
import subprocess
import tempfile
from pathlib import Path
from uuid import UUID
import logging
import openpyxl
from zipfile import ZipFile

from app.services.document.subprocessor.image_processor import image_processor
from app.services.document.subprocessor.table_processor import table_processor
from app.services.document.extraction_tree import ExtractionTree

logger = logging.getLogger(__name__)

def convert_xls_to_xlsx(xls_path: str) -> str:

    xls_path = Path(xls_path)

    with tempfile.TemporaryDirectory() as temp_dir:

        profile_dir = tempfile.mkdtemp(prefix = "lo_profile_")

        try:
            result = subprocess.run(
                [
                    'libreoffice',
                    '--headless',
                    f'-env:UserInstallation=file://{profile_dir}',
                    '--convert-to', 'xlsx',
                    '--outdir', temp_dir,
                    str(xls_path)
                ],
                capture_output = True,
                text = True,
                timeout = 60)
        finally:
            shutil.rmtree(profile_dir, ignore_errors = True)

        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice conversion failed: {result.stderr}")

        converted_path = Path(temp_dir) / f"{xls_path.stem}.xlsx"

        if not converted_path.exists():
            raise RuntimeError(f"Converted file not found: {converted_path}")

        final_path = xls_path.with_suffix('.xlsx')
        final_path.write_bytes(converted_path.read_bytes())

        logger.info(f"Converted {xls_path.name} to {final_path.name}")
        return str(final_path)

class ExcelProcessor:

    def __init__(self):

        # Use global helper processors
        self.image_processor = image_processor
        self.table_processor = table_processor

    def process_sync(
        self,
        file_path: str,
        tree: ExtractionTree | None = None,
        parent_node_id: UUID | None = None) -> ExtractionTree:

        total_images = 0
        total_charts = 0
        total_embedded = 0
        total_media = 0
        converted_file = None

        try:

            # Convert .xls to .xlsx if necessary

            file_path_obj = Path(file_path)
            if file_path_obj.suffix.lower() == '.xls':
                logger.info(f"Converting legacy .xls file: {file_path}")
                converted_file = convert_xls_to_xlsx(file_path)
                file_path = converted_file

            workbook = openpyxl.load_workbook(file_path, data_only = False)
            doc_base_name = Path(file_path).stem

            # Create or use existing tree

            if tree is None:
                tree = ExtractionTree()
                root_id = tree.add_root(
                    source_path = file_path,
                    document_type = "excel",
                    metadata = {})
            else:
                root_id = tree.add_child(
                    parent_id = parent_node_id,
                    source_path = file_path,
                    document_type = "excel",
                    extraction_type = "embedded_document",
                    metadata = {})

            # Process each sheet

            for sheet_idx, sheet_name in enumerate(workbook.sheetnames):

                sheet = workbook[sheet_name]

                # Add sheet node

                sheet_node_id = tree.add_child(
                    parent_id = root_id,
                    source_path = f"{file_path}:sheet:{sheet_name}",
                    document_type = "sheet",
                    extraction_type = "sheet",
                    metadata = {"page_number": sheet_idx})

                # Extract cell data as table

                table_data = []
                for row in sheet.iter_rows():
                    row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                    table_data.append(row_data)

                if table_data:

                    # Format table as markdown

                    table_result = self.table_processor.process(
                        table_data = table_data)

                    if table_result.success:
                        tree.add_child(
                            parent_id = sheet_node_id,
                            source_path = f"{file_path}:sheet:{sheet_name}:data",
                            document_type = "table",
                            extraction_type = "table",
                            metadata = {"output_text": table_result.output_text})

                # Extract images from sheet

                if hasattr(sheet, '_images'):
                    for img_idx, img in enumerate(sheet._images):
                        try:

                            image_bytes = img._data()
                            img_ext = "png"
                            if hasattr(img, 'format'):
                                img_ext = img.format.lower() if img.format else "png"

                            # Run OCR on image bytes

                            output_text = ""
                            try:
                                ocr_result = self.image_processor.process(image_bytes)
                                output_text = ocr_result.output_text if ocr_result.success else ""
                            except Exception:
                                output_text = ""

                            tree.add_child(
                                parent_id = sheet_node_id,
                                source_path = f"{file_path}:sheet:{sheet_name}:image{img_idx}",
                                document_type = "image",
                                extraction_type = "image",
                                metadata = {
                                    "data": base64.b64encode(image_bytes).decode('utf-8'),
                                    "format": img_ext,
                                    "output_text": output_text,
                                    "extraction_location": f"sheet {sheet_name}",
                                    "suggested_filename": f"{doc_base_name}_{sheet_name}_img{img_idx}.{img_ext}"})
                            total_images += 1

                        except Exception as e:
                            logger.debug(f"Image extraction failed in sheet {sheet_name}: {e}")

                # Extract charts from sheet (not supported for now)

                if hasattr(sheet, '_charts'):
                    for chart_idx, chart in enumerate(sheet._charts):
                        try:

                            tree.add_child(
                                parent_id = sheet_node_id,
                                source_path = f"{file_path}:sheet:{sheet_name}:chart{chart_idx}",
                                document_type = "chart",
                                extraction_type = "chart",
                                metadata = {
                                    "chart_type": type(chart).__name__,
                                    "extraction_location": f"sheet {sheet_name}",
                                    "note": "Chart detected but image export not supported"})
                            total_charts += 1

                        except Exception as e:
                            logger.debug(f"Chart extraction failed in sheet {sheet_name}: {e}")

            # Extract embedded files and media from workbook

            try:

                with ZipFile(file_path, 'r') as zf:

                    for name in zf.namelist():

                        if name.startswith('xl/embeddings/') and not name.endswith('/'):
                            
                            try:

                                emb_bytes = zf.read(name)
                                emb_filename = Path(name).name
                                safe_emb_name = f"{doc_base_name}_{emb_filename}"

                                tree.add_child(
                                    parent_id = root_id,
                                    source_path = f"{file_path}:embedded:{emb_filename}",
                                    document_type = "embedded",
                                    extraction_type = "embedded_document",
                                    metadata = {
                                        "data": base64.b64encode(emb_bytes).decode('utf-8'),
                                        "original_filename": emb_filename,
                                        "file_size": len(emb_bytes),
                                        "suggested_filename": safe_emb_name})

                                total_embedded += 1
                                logger.debug(f"Extracted embedded file: {safe_emb_name}")

                            except Exception as e:
                                logger.debug(f"Embedded file extraction failed for {name}: {e}")

                        if name.startswith('xl/media/') and not name.endswith('/'):
                            
                            try:

                                media_filename = Path(name).name
                                ext = Path(media_filename).suffix.lower()

                                if ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.wmf', '.emf']:
                                    continue

                                media_bytes = zf.read(name)
                                safe_media_name = f"{doc_base_name}_{media_filename}"

                                tree.add_child(
                                    parent_id = root_id,
                                    source_path = f"{file_path}:media:{media_filename}",
                                    document_type = "media",
                                    extraction_type = "media",
                                    metadata = {
                                        "data": base64.b64encode(media_bytes).decode('utf-8'),
                                        "original_filename": media_filename,
                                        "file_size": len(media_bytes),
                                        "suggested_filename": safe_media_name})

                                total_media += 1
                                logger.debug(f"Extracted media file: {safe_media_name}")

                            except Exception as e:
                                logger.debug(f"Media file extraction failed for {name}: {e}")

            except Exception as e:
                logger.debug(f"Embedded/media files scan failed: {e}")

            workbook.close()

            logger.info(f"Excel processed: {len(workbook.sheetnames)} sheets, "
                       f"Images: {total_images}, Charts: {total_charts}, "
                       f"Media: {total_media}, Embedded: {total_embedded}")

            return tree

        except Exception as e:
            logger.error(f"Excel processing failed: {e}")
            raise

        finally:

            if converted_file and Path(converted_file).exists():
                try:
                    Path(converted_file).unlink()
                except Exception:
                    pass

# Global instance
excel_processor = ExcelProcessor()
