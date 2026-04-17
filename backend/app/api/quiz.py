from datetime import datetime
import logging
import io
import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from fastapi.responses import StreamingResponse

from app.core.database import get_postgres
from app.core.enums import UserPriority
from app.models.quiz import QuizAttemptCreate, QuizAttemptResponse, EvaluateAnswerRequest, EvaluateAnswerResponse, GenerateMcqChoicesRequest, GenerateMcqChoicesResponse
from app.services.quiz_service import QuizService
from app.services.infrastructure.task_queue_manager import task_queue_manager, QueueType
from app.services.ai.provider import ai_provider, AIProviderError
from app.services.ai.prompts import TEXT_IMPORT_SYSTEM_PROMPT, TEXT_IMPORT_USER_TEMPLATE
from typing import Optional

from app.models.quiz import QuizQuestionResponse

logger = logging.getLogger(__name__)


def _extract_json_from_response(text: str) -> str:
    """Strip markdown code fences and return clean JSON string."""
    match = re.search(r'```(?:json)?\s*([\s\S]*?)\s*```', text)
    if match:
        return match.group(1).strip()
    return text.strip()


router = APIRouter(prefix="/quiz", tags=["Quiz"])


@router.post("/attempts", response_model=QuizAttemptResponse, status_code=status.HTTP_201_CREATED)
async def create_quiz_attempt(payload: QuizAttemptCreate, db=Depends(get_postgres)):
    """Record a quiz attempt into the `quiz_attempts` table."""
    try:
        service = QuizService(db)
        attempt = await service.create_attempt(payload)
        return attempt
    except Exception as e:
        logger.exception("Failed to create quiz attempt")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/attempts")
async def list_quiz_attempts(limit: int = 100, db=Depends(get_postgres)):
    try:
        service = QuizService(db)
        rows = await service.list_attempts(limit=limit)
        return {"attempts": rows}
    except Exception as e:
        logger.exception("Failed to list quiz attempts")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/questions")
async def list_quiz_questions(
    limit: int = 10,
    source_exam: Optional[str] = None,
    difficulty: Optional[int] = None,
    topic: Optional[str] = None,
    db=Depends(get_postgres),
):
    """Return exam questions from `exam_questions` table."""
    try:
        service = QuizService(db)
        rows = await service.list_questions(limit=limit, source_exam=source_exam, difficulty=difficulty, topic=topic)
        return {"questions": rows}
    except Exception as e:
        logger.exception("Failed to list quiz questions")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/topics")
async def list_exam_topics(db=Depends(get_postgres)):
    """Return distinct topics present in the exam_questions table."""
    try:
        rows = await db.fetch(
            "SELECT DISTINCT topic FROM exam_questions WHERE topic IS NOT NULL ORDER BY topic"
        )
        return {"topics": [r["topic"] for r in rows]}
    except Exception as e:
        logger.exception("Failed to list exam topics")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/subjects")
async def list_exam_subjects(db=Depends(get_postgres)):
    """Return all subjects available for exam question filtering."""
    try:
        rows = await db.fetch(
            "SELECT id::text, code, name FROM subjects ORDER BY name"
        )
        return {"subjects": [dict(r) for r in rows]}
    except Exception as e:
        logger.exception("Failed to list exam subjects")
        raise HTTPException(status_code=500, detail=str(e))


def _build_template_workbook() -> openpyxl.Workbook:
    """Build and return the past paper Excel template workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Past Papers"

    headers = [
        "source_exam", "year", "paper", "question_no", "topic",
        "question_stem", "question_type", "options",
        "correct_answer", "answer_explanation", "difficulty_level"
    ]
    required_cols = {"source_exam", "year", "question_stem", "correct_answer"}
    header_fill = PatternFill("solid", fgColor="1F4E79")
    required_fill = PatternFill("solid", fgColor="C00000")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = required_fill if h in required_cols else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    sample_rows = [
        ["DSE", 2024, "P1", "Q1", "Algebra", "Solve for x: 2x + 3 = 11.", "mcq",
         '["A. x=2","B. x=4","C. x=6","D. x=8"]', "B", "2x = 8, so x = 4.", 1],
        ["DSE", 2024, "P1", "Q2", "Calculus", "Differentiate y = 4x^3 with respect to x.", "mcq",
         '["A. 4x^2","B. 12x^2","C. 12x^3","D. 4x^4"]', "B", "Power rule: 3*4x^2 = 12x^2.", 2],
        ["DSE", 2024, "P2", "Q1", "Geometry", "Prove that the sum of angles in a triangle is 180 degrees.", "longq",
         "", "Draw a parallel line through the apex and use alternate interior angles.", "Use parallel line properties.", 3],
        ["ALevel", 2024, "P1", "Q1", "Calculus", "Find dy/dx if y = 3x^2 + 2x - 5.", "mcq",
         '["A. 6x+2","B. 6x-2","C. 3x+2","D. 6x"]', "A", "Differentiate term by term: 6x + 2.", 2],
        ["Mock", 2024, "P1", "Q1", "Trigonometry", "What is sin(30°)?", "mcq",
         '["A. 1/2","B. sqrt(2)/2","C. sqrt(3)/2","D. 1"]', "A", "sin(30°) = 1/2 is a standard value.", 1],
    ]

    even_fill = PatternFill("solid", fgColor="DCE6F1")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    for r, row in enumerate(sample_rows, 2):
        fill = even_fill if r % 2 == 0 else odd_fill
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    col_widths = [12, 6, 6, 10, 14, 45, 12, 45, 14, 45, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 65
    ws2.column_dimensions["C"].width = 12
    for c, h in enumerate(["Column", "Description", "Required?"], 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    instructions = [
        ("source_exam", "Exam type: DSE, ALevel, or Mock", "YES"),
        ("year", "Integer year, e.g. 2024", "YES"),
        ("paper", "Paper number, e.g. P1, P2 (optional)", "no"),
        ("question_no", "Question number, e.g. Q1, Q3b (optional)", "no"),
        ("topic", "Topic name, e.g. Algebra, Calculus (optional)", "no"),
        ("question_stem", "The full question text", "YES"),
        ("question_type", "mcq or longq (default: longq)", "no"),
        ("options", 'JSON array for MCQ: ["A. opt1","B. opt2","C. opt3","D. opt4"] — leave empty for longq', "no"),
        ("correct_answer", "Letter A/B/C/D for MCQ, or full answer text for longq", "YES"),
        ("answer_explanation", "Explanation of the correct answer (optional)", "no"),
        ("difficulty_level", "Integer 1 (easy) to 5 (hard) (optional)", "no"),
    ]
    for r, (col, desc, req) in enumerate(instructions, 2):
        ws2.cell(row=r, column=1, value=col).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=desc).alignment = Alignment(wrap_text=True)
        cell = ws2.cell(row=r, column=3, value=req)
        cell.alignment = Alignment(horizontal="center")
        if req == "YES":
            cell.font = Font(color="C00000", bold=True)
        ws2.row_dimensions[r].height = 25

    return wb


async def _insert_rows(db, rows, start_row=2):
    inserted = 0
    errors = []
    for i, row in enumerate(rows, start=start_row):
        try:
            year_val = row.get("year", 0)
            year = int(str(year_val).strip()) if year_val else 0

            difficulty = str(row.get("difficulty_level", "")).strip()
            difficulty_int = int(difficulty) if difficulty.isdigit() else None

            options_raw = str(row.get("options", "")).strip()
            options_json = None
            if options_raw:
                try:
                    options_json = json.loads(options_raw)
                except json.JSONDecodeError:
                    options_json = None

            await db.execute(
                """
                INSERT INTO exam_questions
                  (source_exam, year, paper, question_no, topic, question_stem,
                   question_type, options, correct_answer, answer_explanation, difficulty_level)
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8::jsonb,$9,$10,$11)
                """,
                str(row.get("source_exam", "")).strip(),
                year,
                str(row.get("paper", "")).strip() or None,
                str(row.get("question_no", "")).strip() or None,
                str(row.get("topic", "")).strip() or None,
                str(row.get("question_stem", "")).strip(),
                str(row.get("question_type", "longq")).strip() or "longq",
                json.dumps(options_json) if options_json else None,
                str(row.get("correct_answer", "")).strip(),
                str(row.get("answer_explanation", "")).strip() or None,
                difficulty_int,
            )
            inserted += 1
        except Exception as e:
            errors.append({"row": i, "error": str(e)})
    return inserted, errors


@router.get("/import-template")
async def download_import_template():
    """Download the Excel template for importing past paper questions."""
    wb = _build_template_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=past_paper_template.xlsx"},
    )


@router.post("/import-csv")
async def import_past_paper_excel(
    file: UploadFile = File(...),
    db=Depends(get_postgres),
):
    """
    Import past paper questions from an Excel (.xlsx) file into exam_questions.

    Required columns: source_exam, year, question_stem, correct_answer
    Optional columns: paper, question_no, topic, question_type, options, answer_explanation, difficulty_level

    - question_type: 'mcq' or 'longq'
    - options: JSON array string for MCQ (e.g. '["A. ...", "B. ..."]') or empty for longq
    - difficulty_level: integer 1-5
    """
    filename = file.filename or ""
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Only Excel (.xlsx) files are accepted")

    required_cols = {"source_exam", "year", "question_stem", "correct_answer"}
    content = await file.read()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        raw_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        if not required_cols.issubset(set(raw_headers)):
            raise HTTPException(
                status_code=400,
                detail=f"Excel must have columns: {', '.join(required_cols)}. Got: {raw_headers}"
            )
        rows = []
        for excel_row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in excel_row):
                continue
            rows.append(dict(zip(raw_headers, [v if v is not None else "" for v in excel_row])))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {e}")

    inserted, errors = await _insert_rows(db, rows)
    return {"inserted": inserted, "errors": errors}


@router.post("/ai-generate")
async def ai_generate_questions(
    topic: str = Form(...),
    source_exam: str = Form("Mock"),
    year: int = Form(...),
    count: int = Form(5),
    difficulty_level: int = Form(3),
    db=Depends(get_postgres),
):
    """
    Use AI to generate exam questions for a given topic and save them
    into exam_questions so they appear in the knowledge base.
    """
    try:
        service = QuizService(db)
        result = await service.ai_generate_questions(
            topic=topic,
            source_exam=source_exam,
            year=year,
            count=min(count, 20),
            difficulty_level=difficulty_level,
        )
        return result
    except Exception as e:
        logger.exception("Failed to AI-generate questions")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/generate-mcq-choices", response_model=GenerateMcqChoicesResponse)
async def generate_mcq_choices(
    payload: GenerateMcqChoicesRequest,
    db=Depends(get_postgres)
):
    """
    Use AI to generate 3 plausible but wrong distractor choices for an open question.
    Returns all 4 choices (correct + 3 distractors) shuffled with the correct answer index.
    """
    try:
        service = QuizService(db)
        result = await service.generate_mcq_choices(payload)
        return result
    except Exception as e:
        logger.exception("Failed to generate MCQ choices")
        raise HTTPException(status_code=500, detail=f"MCQ generation failed: {str(e)}")


@router.post("/import-text")
async def import_text_as_questions(
    text_input: str = Form(...),
    db=Depends(get_postgres),
):
    """Import past paper questions from plain text using Qwen AI to parse and structure them."""
    if not text_input.strip():
        raise HTTPException(status_code=400, detail="Text input cannot be empty")

    try:
        prompt = TEXT_IMPORT_USER_TEMPLATE.format(text_input=text_input)

        async def _ai_import():
            # provider_name=None lets the AI provider try all configured providers
            # in priority order (deepseek → openrouter → qwen → macmini → ollama)
            async with ai_provider.session(
                system_prompt=TEXT_IMPORT_SYSTEM_PROMPT,
                provider_name=None,
            ) as s:
                return await ai_provider.generate(
                    prompt=prompt,
                    session=s,
                    temperature=0.3,
                    max_tokens=8000,
                    user_priority=UserPriority.REGULAR,
                )

        content = await task_queue_manager.submit_and_wait(
            QueueType.AI_GENERATION, _ai_import, UserPriority.REGULAR
        )
        
        try:
            content = _extract_json_from_response(content)
            questions = json.loads(content)
            if not isinstance(questions, list):
                questions = [questions]
        except json.JSONDecodeError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Failed to parse AI response as JSON: {str(e)}\nResponse: {content[:200]}"
            )
        
        inserted = 0
        errors = []
        for i, q in enumerate(questions):
            try:
                if not q.get("question_stem"):
                    errors.append({"index": i, "error": "Missing question_stem"})
                    continue
                
                if not q.get("correct_answer"):
                    errors.append({"index": i, "error": "Missing correct_answer"})
                    continue
                
                source_exam = str(q.get("source_exam", "Mock")).strip()
                year = int(q.get("year", 2024))
                question_type = str(q.get("question_type", "longq")).strip()
                difficulty = int(q.get("difficulty_level", 3))
                
                if source_exam not in ["DSE", "ALevel", "Mock"]:
                    source_exam = "Mock"
                if year < 2000 or year > 2099:
                    year = 2024
                if difficulty < 1 or difficulty > 5:
                    difficulty = 3
                if question_type not in ["mcq", "longq"]:
                    question_type = "longq"
                
                options_raw = q.get("options")
                options_json = None
                if options_raw and isinstance(options_raw, list) and len(options_raw) > 0:
                    options_json = json.dumps(options_raw)
                
                await db.execute(
                    """
                    INSERT INTO exam_questions
                      (source_exam, year, paper, question_no, topic, question_stem,
                       question_type, options, correct_answer, answer_explanation, difficulty_level)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8::jsonb,$9,$10,$11)
                    """,
                    source_exam,
                    year,
                    str(q.get("paper", "")).strip() or None,
                    str(q.get("question_no", "")).strip() or None,
                    str(q.get("topic", "")).strip() or None,
                    str(q.get("question_stem", "")).strip(),
                    question_type,
                    options_json,
                    str(q.get("correct_answer", "")).strip(),
                    str(q.get("answer_explanation", "")).strip() or None,
                    difficulty,
                )
                inserted += 1
            except Exception as e:
                errors.append({"index": i, "error": str(e)})
        
        return {
            "inserted": inserted,
            "errors": errors,
            "message": f"Successfully imported {inserted} questions from text" if inserted > 0 else "No valid questions imported"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Failed to import text as questions")
        raise HTTPException(status_code=500, detail=f"Text import failed: {str(e)}")


@router.post("/evaluate-answer", response_model=EvaluateAnswerResponse)
async def evaluate_answer_with_ai(
    payload: EvaluateAnswerRequest,
    db=Depends(get_postgres)
):
    """
    Evaluate whether a student's answer is correct using AI.
    
    The AI will:
    1. Compare the student's answer against the correct answer
    2. Consider alternative correct wordings
    3. Provide reasoning and constructive feedback
    
    Returns:
    - is_correct: Boolean indicating if answer is correct
    - confidence: Confidence level (0.0 to 1.0)
    - reasoning: Explanation of the evaluation
    - feedback: Constructive feedback for the student
    """
    try:
        service = QuizService(db)
        result = await service.evaluate_answer_with_ai(payload)
        return result
    except ValueError as e:
        if "AI_PARSE_FAILED" in str(e):
            raise HTTPException(status_code=503, detail="AI evaluation temporarily unavailable. Please try again.")
        raise HTTPException(status_code=500, detail=f"Answer evaluation failed: {str(e)}")
    except Exception as e:
        logger.exception("Failed to evaluate answer")
        raise HTTPException(status_code=500, detail=f"Answer evaluation failed: {str(e)}")
